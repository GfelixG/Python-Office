import openpyxl
from openpyxl import Workbook
import datetime
wb = Workbook()

ws = wb.active

ws['A1'] = 42

ws.append([1, 2, 3])


ws['A2'] = datetime.datetime.now()

wb.save("sample.xlsx")

livro = openpyxl.load_workbook('PlanilhasModelo.xlsx')
frutas_pag = book['Frutas']
for rows in frutas_pag.iter_rows(min_rows=2, max_rows=5):
    for cells in rows:
        for cell in rows:
            if cell.value == 'Banana':
                cell.value = 'Fruta 1'
book.save('PlanilhasModelo v2.xlsx')
